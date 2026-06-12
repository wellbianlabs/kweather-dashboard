"""리포트 자동 생성 (PRD 3.4).

- PDF: (있으면) matplotlib 차트(서버 렌더) + Jinja2 HTML + xhtml2pdf (A4, page-break)
- Excel: openpyxl (요약 + 로우데이터, 정렬된 .xlsx)
한글 폰트는 번들된 NanumGothic 우선, 없으면 Windows '맑은 고딕'.

서버리스(Vercel) 용량 한도를 위해 matplotlib 는 선택적 의존성으로 처리한다.
설치되어 있으면 PDF 에 차트 이미지를 임베드하고, 없으면 표·가이드만 출력한다.
"""
from __future__ import annotations

import base64
import io
import os
from datetime import date as date_cls, datetime, time

import pandas as pd
from jinja2 import Template
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session
from xhtml2pdf import pisa

from .. import heat
from ..config import settings
from ..models import Device, ExternalDailyCache, Tenant
from . import analytics

# ---- matplotlib (선택적): 없으면 차트 이미지 생략 ----
try:
    import matplotlib

    matplotlib.use("Agg")  # GUI 없는 서버 렌더
    import matplotlib.dates as mdates
    import matplotlib.pyplot as plt
    from matplotlib import font_manager

    HAS_MPL = True
except Exception:  # noqa: BLE001
    HAS_MPL = False

# ---- 한글 폰트 등록 ----
from reportlab.pdfbase import pdfmetrics  # noqa: E402
from reportlab.pdfbase.ttfonts import TTFont as RLTTFont  # noqa: E402
from xhtml2pdf.default import DEFAULT_FONT  # noqa: E402

_BUNDLED_FONT = os.path.join(os.path.dirname(__file__), "..", "fonts", "NanumGothic-Regular.ttf")
_FONT_CANDIDATES = [
    os.path.normpath(_BUNDLED_FONT),  # 배포(Linux)/로컬 공통 — 저장소에 동봉
    r"C:\Windows\Fonts\malgun.ttf",
    r"C:\Windows\Fonts\malgunsl.ttf",
]
_FONT_PATH = next((p for p in _FONT_CANDIDATES if os.path.exists(p)), None)
_PDF_FONT = "Helvetica"  # 폴백
if _FONT_PATH:
    if HAS_MPL:
        font_manager.fontManager.addfont(_FONT_PATH)
        plt.rcParams["font.family"] = font_manager.FontProperties(fname=_FONT_PATH).get_name()
    # reportlab 에 폰트 등록 + 패밀리(굵게/기울임도 동일 한글 폰트로) 매핑.
    # @font-face 는 Windows 에서 임시파일 잠금 버그가 있으므로 사용하지 않고,
    # xhtml2pdf 의 폰트 매핑 테이블(DEFAULT_FONT)에 직접 등록해 font-family 를 해석시킨다.
    try:
        pdfmetrics.registerFont(RLTTFont("KFont", _FONT_PATH))
        pdfmetrics.registerFontFamily(
            "KFont", normal="KFont", bold="KFont", italic="KFont", boldItalic="KFont"
        )
        DEFAULT_FONT["kfont"] = "KFont"
        _PDF_FONT = "KFont"
    except Exception:  # noqa: BLE001
        _PDF_FONT = "Helvetica"
if HAS_MPL:
    plt.rcParams["axes.unicode_minus"] = False


def _fig_to_data_uri(fig) -> str:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight")
    plt.close(fig)
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")
    return f"data:image/png;base64,{b64}"


def _daily_chart(db: Session, tenant: Tenant, device_sn: str, on_date: date_cls) -> str | None:
    if not HAS_MPL:
        return None
    start = datetime.combine(on_date, time.min)
    end = datetime.combine(on_date, time.max)
    ts = analytics.time_series(db, tenant, device_sn, start, end, 10)
    if not ts.points:
        return None
    xs = [p.t for p in ts.points]
    feels = [p.feels_like for p in ts.points]
    temps = [p.temperature for p in ts.points]

    fig, ax = plt.subplots(figsize=(9, 3.6))
    ax.plot(xs, feels, color="#dc2626", linewidth=2, label="체감온도(A-TEMP)")
    ax.plot(xs, temps, color="#2563eb", linewidth=1.2, alpha=0.7, label="온도(TEMP)")
    for key, lvl in (("attention", "관심"), ("caution", "주의"), ("warning", "경고"), ("danger", "위험")):
        ax.axhline(heat.thresholds()[key], color=heat.LEVELS[key].color, linestyle="--", linewidth=0.9, alpha=0.7)
    ax.set_ylabel("온도 (°C)")
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    ax.legend(loc="upper left", fontsize=8)
    ax.grid(True, alpha=0.25)
    ax.set_title(f"{device_sn} — {on_date.isoformat()} 체감온도 추이")
    return _fig_to_data_uri(fig)


def _daily_detail(db: Session, tenant: Tenant, device_sn: str, on_date: date_cls) -> dict:
    """일일 상세 리포트용 데이터 — KPI, 단계별 지속시간, 시간대별 집계, 내부 vs 외부(기상청) 비교, 분석 코멘트."""
    from . import weather as weather_svc  # 지연 임포트(순환 방지)

    start = datetime.combine(on_date, time.min)
    end = datetime.combine(on_date, time.max)
    dev = db.get(Device, device_sn)
    th = heat.thresholds()
    df = analytics.load_logs(db, [device_sn], start, end)

    out: dict = {
        "device_sn": device_sn,
        "company_name": dev.company_name if dev else None,
        "location_name": dev.location_name if dev else None,
        "address": dev.address if dev else None,
        "date": on_date.isoformat(),
        "levels": heat.LEVELS,
        "has_data": not df.empty,
    }
    if df.empty:
        safe = heat.LEVELS["safe"]
        out.update(
            peak_label=safe.label, peak_color=safe.color, guidance=analytics._GUIDANCE["safe"],
            hours=[], level_minutes={}, total_minutes=0, weather=None, analysis=[],
            external_daily=None, avg_humidity=None, work=None, series=[],
        )
        return out

    feels, temps, humi = df["feels_like"], df["temperature"], df["humidity"]
    n = len(df)
    idx_max = feels.idxmax()
    max_feels = round(float(feels.loc[idx_max]), 1)
    max_time = pd.to_datetime(df.loc[idx_max, "measured_at"]).strftime("%H:%M")
    peak = heat.classify(max_feels)

    # 단계별 누적 분 (1분 주기 가정)
    lm = {
        "danger": int((feels >= th["danger"]).sum()),
        "warning": int(((feels >= th["warning"]) & (feels < th["danger"])).sum()),
        "caution": int(((feels >= th["caution"]) & (feels < th["warning"])).sum()),
        "attention": int(((feels >= th["attention"]) & (feels < th["caution"])).sum()),
        "safe": int((feels < th["attention"]).sum()),
    }

    # 피크 시점의 동시 관측값(샘플 보고서 항목)
    temp_at_peak = round(float(df.loc[idx_max, "temperature"]), 1)
    _hp = df.loc[idx_max, "humidity"]
    humi_at_peak = int(_hp) if pd.notna(_hp) else None

    # 차트용 10분 시리즈 (시각을 0~24h 실수로)
    s10 = df.set_index("measured_at")["feels_like"].resample("10min").mean().dropna()
    series = [(ts.hour + ts.minute / 60.0, round(float(v), 1)) for ts, v in s10.items()]

    # 시간대별 평균
    s = df.set_index("measured_at")[["temperature", "feels_like", "humidity"]].resample("1h").mean()

    # ---- 외부 시간자료(측정 당시의 기상청 기온·습도·공식 체감온도) — 캐시 우선 ----
    import json as _json

    ds_key = on_date.strftime("%Y%m%d")
    ext_hourly: dict[int, dict] | None = None
    cache_row = db.scalar(
        select(ExternalDailyCache).where(
            ExternalDailyCache.device_sn == device_sn, ExternalDailyCache.ymd == ds_key
        )
    )
    if cache_row and cache_row.hourly_json:
        try:
            ext_hourly = {int(k): v for k, v in _json.loads(cache_row.hourly_json).items()}
        except Exception:  # noqa: BLE001
            ext_hourly = None
    if ext_hourly is None:
        code = weather_svc.resolve_dong_code(dev)
        fetched = weather_svc._kma_asos_hourly(code, ds_key) if code else None
        if fetched:
            ext_hourly = fetched
            try:
                if cache_row is None:
                    cache_row = ExternalDailyCache(device_sn=device_sn, ymd=ds_key)
                    db.add(cache_row)
                cache_row.hourly_json = _json.dumps({str(k): v for k, v in fetched.items()})
                db.commit()
            except Exception:  # noqa: BLE001
                db.rollback()

    # 실시간(provider) 비교 — 외부 시간자료가 없을 때의 폴백 소스
    try:
        cmp = weather_svc.compare(db, tenant, device_sn, start, end, 60)
    except Exception:  # noqa: BLE001
        cmp = None
    out_by_h = {pd.Timestamp(p.t).hour: p.outdoor_temperature for p in cmp.points} if cmp else {}

    hours = []
    for idx, row in s.iterrows():
        h = idx.hour
        f = None if pd.isna(row["feels_like"]) else round(float(row["feels_like"]), 1)
        lvl = heat.classify(f)
        slot = ext_hourly.get(h) if ext_hourly else None
        o_ta = slot.get("ta") if slot else out_by_h.get(h)
        o_fl = slot.get("feels") if slot else None
        base = o_fl if o_fl is not None else o_ta
        delta = round(f - base, 1) if (f is not None and base is not None) else None
        hours.append({
            "hour": h, "feels": f,
            "temp": None if pd.isna(row["temperature"]) else round(float(row["temperature"]), 1),
            "humidity": None if pd.isna(row["humidity"]) else int(round(float(row["humidity"]))),
            "label": lvl.label, "color": lvl.color,
            "outdoor": o_ta, "out_feels": o_fl, "delta": delta,
        })

    deltas = [x["delta"] for x in hours if x["delta"] is not None]
    avg_delta = round(sum(deltas) / len(deltas), 1) if deltas else None
    avg_humi = round(float(humi.mean()), 1) if humi.notna().any() else None
    has_out_feels = any(x["out_feels"] is not None for x in hours)
    out_feels_vals = [x["out_feels"] for x in hours if x["out_feels"] is not None]
    out_feels_max = round(max(out_feels_vals), 1) if out_feels_vals else None
    out_feels_avg = round(sum(out_feels_vals) / len(out_feels_vals), 1) if out_feels_vals else None

    # 근무시간(09:00~18:00) 통계 — 근로자 보호 관점의 핵심 구간
    hrs = df["measured_at"].dt.hour
    wdf = df[(hrs >= 9) & (hrs < 18)]
    work = None
    if not wdf.empty:
        wfeels = wdf["feels_like"]
        widx = wfeels.idxmax()
        wpeak = heat.classify(round(float(wfeels.max()), 1))
        work = {
            "max_feels": round(float(wfeels.max()), 1),
            "max_time": pd.to_datetime(wdf.loc[widx, "measured_at"]).strftime("%H:%M"),
            "max_temp": round(float(wdf["temperature"].max()), 1),
            "avg_feels": round(float(wfeels.mean()), 1),
            "danger_minutes": int((wfeels >= th["danger"]).sum()),
            "minutes": {
                "danger": int((wfeels >= th["danger"]).sum()),
                "warning": int(((wfeels >= th["warning"]) & (wfeels < th["danger"])).sum()),
                "caution": int(((wfeels >= th["caution"]) & (wfeels < th["warning"])).sum()),
                "attention": int(((wfeels >= th["attention"]) & (wfeels < th["caution"])).sum()),
                "safe": int((wfeels < th["attention"]).sum()),
            },
            "total": len(wdf),
            "peak_label": wpeak.label, "peak_color": wpeak.color,
        }

    weather = None
    if deltas:
        max_delta = round(max(deltas), 1)
        weather = {
            "provider": "kma" if has_out_feels else (cmp.provider if cmp else "mock"),
            "max_delta": max_delta, "avg_delta": avg_delta,
            "enclosed_alert": max_delta >= settings.ENCLOSED_DELTA_ALERT,
            "threshold": settings.ENCLOSED_DELTA_ALERT,
            "feels_based": has_out_feels,
        }

    # 외부 일별 요약(과거자료): 캐시 우선 → 아카이브/ASOS 일자료 → 시간자료 집계 폴백
    external_daily = None
    provider = weather_svc.get_provider()
    ed = None
    if cache_row and (cache_row.max_temp is not None or cache_row.avg_temp is not None):
        ed = {
            "avg": float(cache_row.avg_temp) if cache_row.avg_temp is not None else None,
            "max": float(cache_row.max_temp) if cache_row.max_temp is not None else None,
            "min": float(cache_row.min_temp) if cache_row.min_temp is not None else None,
            "humi": float(cache_row.humidity) if cache_row.humidity is not None else None,
            "source": cache_row.source, "region": cache_row.region,
        }
    else:
        if hasattr(provider, "past_daily"):
            try:
                ed = provider.past_daily(dev.latitude, dev.longitude, dev.region_code, on_date)
            except Exception:  # noqa: BLE001
                ed = None
        # 일별 아카이브가 없으면 시간자료에서 직접 집계 (측정 당시 기준)
        if (not ed or ed.get("max") is None) and ext_hourly:
            tas = [v["ta"] for v in ext_hourly.values() if v.get("ta") is not None]
            hms = [v["hm"] for v in ext_hourly.values() if v.get("hm") is not None]
            if tas:
                ed = {
                    "avg": round(sum(tas) / len(tas), 1), "max": round(max(tas), 1),
                    "min": round(min(tas), 1),
                    "humi": round(sum(hms) / len(hms), 1) if hms else None,
                    "source": "케이웨더 기상관측자료", "region": None,
                }
        if ed and (ed.get("avg") is not None or ed.get("max") is not None):
            try:
                if cache_row is None:
                    cache_row = ExternalDailyCache(device_sn=device_sn, ymd=ds_key)
                    db.add(cache_row)
                cache_row.avg_temp = ed.get("avg"); cache_row.max_temp = ed.get("max")
                cache_row.min_temp = ed.get("min"); cache_row.humidity = ed.get("humi")
                cache_row.source = ed.get("source"); cache_row.region = ed.get("region")
                db.commit()
            except Exception:  # noqa: BLE001
                db.rollback()

    if ed and (ed.get("avg") is not None or ed.get("max") is not None):
        in_max = round(float(temps.max()), 1)
        in_avg = round(float(temps.mean()), 1)
        external_daily = {
            "region": ed.get("region"), "source": ed.get("source"),
            "out_avg": ed.get("avg"), "out_max": ed.get("max"), "out_min": ed.get("min"), "out_humi": ed.get("humi"),
            "out_feels_max": out_feels_max, "out_feels_avg": out_feels_avg,
            "in_avg": in_avg, "in_max": in_max,
            "diff_max": round(in_max - float(ed["max"]), 1) if ed.get("max") is not None else None,
            "diff_feels": round(max_feels - out_feels_max, 1) if out_feels_max is not None else None,
        }

    # 자동 분석 코멘트
    analysis: list[str] = []
    if work:
        analysis.append(
            f"근무시간(09:00~18:00) 중 최고 체감온도는 {work['max_time']}경 {work['max_feels']}°C(단계: {work['peak_label']})이며, "
            f"위험단계(38°C 이상) 노출이 {work['danger_minutes']}분 누적됨."
        )
    if lm["danger"]:
        analysis.append(f"체감온도 38°C(위험) 이상 노출이 일일 {lm['danger']}분 누적되어 고용노동부 기준상 옥외작업 원칙적 중지 대상에 해당함.")
    analysis.append(f"최고 체감온도는 {max_time}경 {max_feels}°C로 관측되어 일중 최고치를 기록함"
                    + (f" (당시 기온 {temp_at_peak}°C, 습도 {humi_at_peak}%)." if humi_at_peak is not None else f" (당시 기온 {temp_at_peak}°C)."))
    base_label = "공식 체감온도" if has_out_feels else "기온"
    src_label = "케이웨더" if has_out_feels or (cmp and cmp.provider in ("kweather", "kma")) else "참고용 추정"
    if weather and weather["enclosed_alert"]:
        analysis.append(
            f"작업장 내부 체감온도가 외부({src_label}) {base_label} 대비 최대 {weather['max_delta']}°C 높게 측정되어 "
            f"'밀폐형 폭염 사업장'에 해당함(관리 임계 {weather['threshold']}°C 초과). 환기·차열·국소냉방 등 작업환경 개선 필요."
        )
    elif weather and avg_delta is not None:
        analysis.append(f"작업장 내부 체감온도가 외부({src_label}) {base_label} 대비 평균 {avg_delta}°C 높게 측정됨(최대 {weather['max_delta']}°C).")

    if avg_humi is not None and avg_humi >= 70:
        analysis.append(f"평균 습도 {avg_humi}%의 고온다습 환경으로 체열 발산이 저해되어 온열질환 발생 위험이 가중되는 조건임.")

    out.update(
        max_feels=max_feels, max_time=max_time, max_temp=round(float(temps.max()), 1),
        avg_feels=round(float(feels.mean()), 1), avg_humidity=avg_humi, record_count=n,
        range_start=pd.to_datetime(df["measured_at"].min()).strftime("%H:%M"),
        range_end=pd.to_datetime(df["measured_at"].max()).strftime("%H:%M"),
        peak_label=peak.label, peak_color=peak.color, guidance=analytics._GUIDANCE[peak.code],
        level_minutes=lm, total_minutes=n, hours=hours, weather=weather, analysis=analysis,
        external_daily=external_daily, work=work, series=series,
        temp_at_peak=temp_at_peak, humi_at_peak=humi_at_peak,
    )
    if external_daily and external_daily.get("out_max") is not None and external_daily.get("diff_feels") is None:
        analysis.append(
            f"외부({external_daily['source']}) 일 최고기온 {external_daily['out_max']}°C 대비 작업장 최고기온 "
            f"{external_daily['in_max']}°C로 {external_daily['diff_max']}°C 편차를 보임."
        )
    return out



# ---------------- PIL 경량 차트 (matplotlib 없이 — 로컬/서버리스 동일 출력) ----------------
def _png_data_uri(img) -> str:
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode("ascii")


def _pil_fonts():
    from PIL import ImageFont

    def F(size):
        try:
            return ImageFont.truetype(_FONT_PATH, size)
        except Exception:  # noqa: BLE001
            return ImageFont.load_default()
    return F


def _chart_hourly_feels(series, th) -> str | None:
    """시간별 체감온도 라인 차트 — 위험단계 색상 구간선 + 임계선 + 피크 주석."""
    try:
        from PIL import Image, ImageDraw
    except Exception:  # noqa: BLE001
        return None
    if not series or len(series) < 2:
        return None
    F = _pil_fonts()
    W, H = 1560, 345
    L, R, T, B = 100, 36, 24, 52
    img = Image.new("RGB", (W, H), "white")
    d = ImageDraw.Draw(img)

    ys = [v for _, v in series]
    ymin = min(min(ys) - 2, 20)
    ymax = max(max(ys) + 3, 41)
    ymin = int(ymin // 5 * 5)
    ymax = int(-(-ymax // 5) * 5)

    def X(x):
        return L + (x / 24.0) * (W - L - R)

    def Y(y):
        return T + (1 - (y - ymin) / (ymax - ymin)) * (H - T - B)

    # 그리드/축
    for gy in range(ymin, ymax + 1, 5):
        d.line([(L, Y(gy)), (W - R, Y(gy))], fill="#eef2f6", width=2)
        d.text((L - 14, Y(gy)), str(gy), font=F(24), fill="#94a3b8", anchor="rm")
    for gx in range(0, 25, 3):
        d.line([(X(gx), T), (X(gx), H - B)], fill="#f4f6f9", width=2)
        d.text((X(gx), H - B + 12), f"{gx:02d}시", font=F(24), fill="#94a3b8", anchor="ma")
    d.line([(L, H - B), (W - R, H - B)], fill="#cbd5e1", width=3)
    d.line([(L, T), (L, H - B)], fill="#cbd5e1", width=3)

    # 임계선(점선)
    for code in ("attention", "caution", "warning", "danger"):
        yv = th[code]
        if ymin < yv < ymax:
            color = heat.LEVELS[code].color
            x = L
            while x < W - R:
                d.line([(x, Y(yv)), (min(x + 16, W - R), Y(yv))], fill=color, width=2)
                x += 28
            d.text((W - R - 4, Y(yv) - 4), f"{heat.LEVELS[code].label} {int(yv)}", font=F(20), fill=color, anchor="rs")

    # 근무시간 음영(09~18시)
    band = Image.new("RGBA", (int(X(18)) - int(X(9)), int(H - B - T)), (15, 73, 158, 14))
    img.paste(band, (int(X(9)), int(T)), band)

    # 단계 색상 구간 폴리라인
    for i in range(len(series) - 1):
        (x1, v1), (x2, v2) = series[i], series[i + 1]
        seg_color = heat.classify((v1 + v2) / 2).color
        d.line([(X(x1), Y(v1)), (X(x2), Y(v2))], fill=seg_color, width=5)

    # 피크 주석
    pi = max(range(len(series)), key=lambda i: series[i][1])
    px_, pv = series[pi]
    pc = heat.classify(pv).color
    d.ellipse([X(px_) - 9, Y(pv) - 9, X(px_) + 9, Y(pv) + 9], fill="white", outline=pc, width=4)
    d.text((X(px_), Y(pv) - 18), f"{pv:.1f}", font=F(30), fill=pc, anchor="mb", stroke_width=1, stroke_fill=pc)

    return _png_data_uri(img)


def _chart_compare(hours) -> str | None:
    """내부 체감온도 vs 야외 기온 비교 라인 차트 (시간 단위)."""
    try:
        from PIL import Image, ImageDraw
    except Exception:  # noqa: BLE001
        return None
    pts_in = [(h["hour"], h["feels"]) for h in hours if h.get("feels") is not None]
    use_feels = sum(1 for h in hours if h.get("out_feels") is not None) >= 2
    key = "out_feels" if use_feels else "outdoor"
    pts_out = [(h["hour"], h[key]) for h in hours if h.get(key) is not None]
    out_label = "야외 체감온도(기상청 공식)" if use_feels else "야외 기온"
    if len(pts_in) < 2 or len(pts_out) < 2:
        return None
    F = _pil_fonts()
    W, H = 1560, 320
    L, R, T, B = 100, 36, 40, 52
    img = Image.new("RGB", (W, H), "white")
    d = ImageDraw.Draw(img)

    ys = [v for _, v in pts_in] + [v for _, v in pts_out]
    ymin = int((min(ys) - 2) // 5 * 5)
    ymax = int(-(-(max(ys) + 3) // 5) * 5)

    def X(x):
        return L + (x / 24.0) * (W - L - R)

    def Y(y):
        return T + (1 - (y - ymin) / (ymax - ymin)) * (H - T - B)

    for gy in range(ymin, ymax + 1, 5):
        d.line([(L, Y(gy)), (W - R, Y(gy))], fill="#eef2f6", width=2)
        d.text((L - 14, Y(gy)), str(gy), font=F(24), fill="#94a3b8", anchor="rm")
    for gx in range(0, 25, 3):
        d.text((X(gx), H - B + 12), f"{gx:02d}시", font=F(24), fill="#94a3b8", anchor="ma")
    d.line([(L, H - B), (W - R, H - B)], fill="#cbd5e1", width=3)
    d.line([(L, T), (L, H - B)], fill="#cbd5e1", width=3)

    def poly(pts, color):
        for i in range(len(pts) - 1):
            d.line([(X(pts[i][0]), Y(pts[i][1])), (X(pts[i + 1][0]), Y(pts[i + 1][1]))], fill=color, width=5)
        for x, v in pts:
            d.ellipse([X(x) - 4, Y(v) - 4, X(x) + 4, Y(v) + 4], fill=color)

    poly(pts_out, "#1790cd")
    poly(pts_in, "#dc2626")

    # 범례
    lx = W - R - 430
    d.line([(lx - 120, 28), (lx - 76, 28)], fill="#dc2626", width=6)
    d.text((lx - 66, 28), "현장 체감온도", font=F(24), fill="#334155", anchor="lm")
    d.line([(lx + 120, 28), (lx + 164, 28)], fill="#1790cd", width=6)
    d.text((lx + 174, 28), out_label, font=F(24), fill="#334155", anchor="lm")

    return _png_data_uri(img)


_DAILY_TEMPLATE = Template(
    """
<html><head><style>
@page { size: A4; margin: 1.5cm 1.6cm; }
body { font-family: "{{ pdf_font }}"; font-size: 9pt; color:#1f2937; line-height:1.5; }
.title { text-align:center; font-size:16pt; font-weight:bold; color:#0f172a; margin:0 0 3pt 0; }
table { width:100%; border-collapse: collapse; }

/* 문서정보 */
.subtitle { text-align:center; font-size:9pt; color:#64748b; margin:0 0 6pt 0; padding-bottom:6pt; border-bottom:1.5px solid #0f499e; }
.docinfo td { border:1px solid #cbd5e1; padding:4px 8px; font-size:8.5pt; }
.docinfo .k { background:#f8fafc; color:#475569; width:14%; text-align:center; }

/* 섹션 */
h2 { font-size:11pt; color:#0f172a; margin:8pt 0 3pt 0; }
h2 .no { color:#0f499e; }
.tbl th { border:1px solid #94a3b8; background:#eef2f7; padding:4px 6px; font-size:8.5pt; color:#334155; text-align:center; }
.tbl td { border:1px solid #cbd5e1; padding:3px 5px; font-size:8.8pt; text-align:center; }
.tbl .k { background:#f8fafc; color:#475569; text-align:center; }
.num { font-weight:bold; font-size:10pt; }
.badge { display:inline-block; padding:1.5px 8px; border-radius:8px; color:#fff; font-weight:bold; font-size:8.5pt; }
.strip { table-layout:fixed; }
.h24 { table-layout:fixed; }
.h24 td { border:1px solid #fff; padding:2px 0; text-align:center; font-size:5.6pt; line-height:1.25; }
.h24 .k { background:#f1f5f9; color:#475569; font-size:6.2pt; }
.strip td { border:1px solid #fff; padding:2.5px 0; text-align:center; color:#fff; font-size:6pt; line-height:1.2; }
.alert { border:1px solid #fca5a5; background:#fef2f2; color:#b91c1c; padding:4px 7px; font-size:8.5pt; margin:3px 0; }
.gov { margin:2pt 0 0 0; }
.gov div { margin:1pt 0; font-size:9pt; }
.gov .b { color:#0f499e; font-weight:bold; }
.gov2 { margin:2pt 0 0 8pt; }
.gov2 div { margin:1pt 0; font-size:9pt; }
.gov2 .b { color:#334155; }
.note { font-size:7.8pt; color:#64748b; margin:1pt 0; }
.footer { margin-top:6pt; border-top:1.5px solid #0f499e; padding-top:4pt; font-size:7pt; color:#64748b; line-height:1.45; }
</style></head><body>

<div class="title">폭염 안전관리 일일 보고서</div>
<div style="text-align:center; font-size:8pt; color:#94a3b8; letter-spacing:1.5pt; margin-bottom:2pt;">HEAT STRESS DAILY MANAGEMENT REPORT</div>
<div class="subtitle">근로자 온열질환 예방을 위한 작업장 체감온도 분석 자료 · 측정장비: 케이웨더(주) 체감온도계</div>

<table class="docinfo">
  <tr>
    <td class="k">보고서 번호</td><td style="width:36%">{{ report_no }}</td>
    <td class="k">작성 일시</td><td>{{ generated }}</td>
  </tr>
  <tr>
    <td class="k">대상 일자</td><td>{{ d.date }}{% if d.has_data %} ({{ d.range_start }}~{{ d.range_end }}, 총 {{ d.record_count }}건 측정){% endif %}</td>
    <td class="k">최고 위험단계</td><td><span class="badge" style="background:{{ d.peak_color }}">{{ d.peak_label }}</span></td>
  </tr>
</table>

<h2><span class="no">1.</span> 측정 대상 개요</h2>
<table class="tbl">
  <tr><th style="width:14%">사업장</th><td style="width:36%">{{ d.company_name or '-' }}</td>
      <th style="width:14%">설치 위치</th><td>{{ d.location_name or '-' }}</td></tr>
  <tr><th>소재지</th><td>{{ d.address or '-' }}</td>
      <th>측정기기</th><td>케이웨더(주) 체감온도계 (SN: {{ d.device_sn }})</td></tr>
</table>
<p class="note">※ 본 보고서의 모든 측정 데이터는 <b>케이웨더(주) 체감온도계 장비</b>로 측정·수집된 자료임.</p>

{% if d.has_data %}
<h2><span class="no">2.</span> 측정 결과 요약 <span style="font-size:8pt; color:#64748b; font-weight:normal;">(근무시간: 09:00~18:00)</span></h2>
<table class="tbl">
  <tr><th style="width:17%">구분</th><th>최고 체감온도</th><th>발생 시각</th><th>최고 기온</th><th>평균 체감온도</th><th>위험단계(38°C↑) 노출</th></tr>
  {% if d.work %}
  <tr style="background:#fbfdff;">
    <td class="k"><b>근무시간</b></td>
    <td class="num" style="color:{{ d.work.peak_color }}">{{ d.work.max_feels }}°C</td>
    <td>{{ d.work.max_time }}</td>
    <td class="num">{{ d.work.max_temp }}°C</td>
    <td>{{ d.work.avg_feels }}°C</td>
    <td class="num" style="color:#dc2626">{{ d.work.danger_minutes }}분</td>
  </tr>
  {% endif %}
  <tr>
    <td class="k">전일(24시간)</td>
    <td class="num" style="color:{{ d.peak_color }}">{{ d.max_feels }}°C</td>
    <td>{{ d.max_time }}</td>
    <td class="num">{{ d.max_temp }}°C</td>
    <td>{{ d.avg_feels }}°C</td>
    <td class="num" style="color:#dc2626">{{ d.level_minutes['danger'] }}분</td>
  </tr>
</table>
<p class="note">※ 평균 습도(전일): {{ d.avg_humidity if d.avg_humidity is not none else '-' }}% · 근로자 보호 관점에서 근무시간(09~18시) 수치를 우선 검토</p>

<h2><span class="no">3.</span> 폭염 위험단계별 노출시간 분석</h2>
<table class="tbl">
  <tr><th style="width:14%">위험 단계</th>{% for code in ['safe','attention','caution','warning','danger'] %}<th style="background:{{ d.levels[code].color }}; color:#fff;">{{ d.levels[code].label }}</th>{% endfor %}</tr>
  <tr><td class="k">기준(체감)</td><td>31°C 미만</td><td>31°C 이상</td><td>33°C 이상</td><td>35°C 이상</td><td>38°C 이상</td></tr>
  {% if d.work %}<tr style="background:#fbfdff;"><td class="k"><b>근무시간 노출</b></td>{% for code in ['safe','attention','caution','warning','danger'] %}<td><b>{{ d.work.minutes[code] }}분</b></td>{% endfor %}</tr>{% endif %}
  <tr><td class="k">전일 노출</td>{% for code in ['safe','attention','caution','warning','danger'] %}<td>{{ d.level_minutes[code] }}분</td>{% endfor %}</tr>
  <tr><td class="k">전일 비율</td>{% for code in ['safe','attention','caution','warning','danger'] %}<td>{{ ((d.level_minutes[code] / d.total_minutes * 100) | round(1)) if d.total_minutes else 0 }}%</td>{% endfor %}</tr>
</table>
<p class="note">※ 측정주기(1분) 기준 누적 노출시간 · 근무시간 = 09:00~18:00 · 단계 기준: 고용노동부 폭염 단계별 대응요령(체감온도)</p>

<h2><span class="no">4.</span> 시간별 체감온도 변화 <span style="font-size:8pt; color:#64748b; font-weight:normal;">(전일 24시간 · 음영구간 = 근무시간 09~18시)</span></h2>
{% if d.hours %}
<table class="h24">
  <tr><td class="k" style="width:34pt;">시각</td>{% for h in d.hours %}<td class="k">{{ '%02d'|format(h.hour) }}</td>{% endfor %}</tr>
  <tr><td class="k">체감(°C)</td>{% for h in d.hours %}<td style="background:{{ h.color }}; color:#fff; font-weight:bold;">{{ h.feels if h.feels is not none else '-' }}</td>{% endfor %}</tr>
</table>
{% endif %}
{% if chart %}<div style="margin-top:6pt;"><img src="{{ chart }}" style="width:480pt;"/></div>{% endif %}
<p class="note">※ 표 색상은 시간대 평균 체감온도의 폭염 위험단계 · 그래프 점선은 단계 임계값, 음영 구간은 근무시간(09:00~18:00)</p>

<h2><span class="no">5.</span> 내·외부 기온 비교 분석 <span style="font-size:8pt; color:#64748b; font-weight:normal;">(근무시간 기준 · 외부: 케이웨더 기상관측자료)</span></h2>
{% if d.external_daily %}
  <table class="tbl" style="margin-bottom:4pt;">
    <tr><th style="width:20%">구분</th><th>최고 체감온도</th><th>평균 체감온도</th><th>일 최고기온</th><th>일 평균기온</th><th>평균 습도</th></tr>
    <tr><td class="k">외부 · 기상청 공식</td>
        <td class="num" style="color:#1790cd;">{{ d.external_daily.out_feels_max if d.external_daily.out_feels_max is not none else '-' }}°C</td>
        <td>{{ d.external_daily.out_feels_avg if d.external_daily.out_feels_avg is not none else '-' }}°C</td>
        <td>{{ d.external_daily.out_max if d.external_daily.out_max is not none else '-' }}°C</td>
        <td>{{ d.external_daily.out_avg if d.external_daily.out_avg is not none else '-' }}°C</td>
        <td>{{ d.external_daily.out_humi if d.external_daily.out_humi is not none else '-' }}%</td></tr>
    <tr><td class="k">작업장(내부 측정)</td>
        <td class="num" style="color:#dc2626;">{{ d.max_feels }}°C</td>
        <td>{{ d.avg_feels }}°C</td>
        <td>{{ d.external_daily.in_max }}°C</td>
        <td>{{ d.external_daily.in_avg }}°C</td>
        <td>{{ d.avg_humidity if d.avg_humidity is not none else '-' }}%</td></tr>
    {% if d.external_daily.diff_feels is not none %}
    <tr><td class="k">체감온도 차(내-외)</td>
        <td class="num" style="color:#b91c1c;">+{{ d.external_daily.diff_feels }}°C</td>
        <td colspan="4" style="text-align:left; font-size:8pt; color:#64748b;">작업장 체감온도가 기상청 공식 외부 체감온도보다 높을수록 복사열·밀폐 영향이 큼</td></tr>
    {% endif %}
  </table>
  <p class="note">※ 출처: {{ d.external_daily.source }} · 작업장 최고기온이 외부 일 최고기온 대비 {{ d.external_daily.diff_max }}°C {{ '높음' if (d.external_daily.diff_max or 0) >= 0 else '낮음' }} (복사열·환기 영향 지표)</p>
{% endif %}
{% if d.weather %}
  {% if d.weather.enclosed_alert %}
  <div class="alert"><b>[경고] 밀폐형 폭염 사업장</b> — 내부 체감온도가 외부 {{ '공식 체감온도' if d.weather.feels_based else '기온' }} 대비 최대 {{ d.weather.max_delta }}°C, 평균 {{ d.weather.avg_delta }}°C 높게 측정됨(관리 임계 {{ d.weather.threshold }}°C 초과). 환기·차열·국소냉방 등 작업환경 개선 필요.</div>
  {% endif %}
  {% if chart2 %}<div style="margin:2pt 0 6pt 0;"><img src="{{ chart2 }}" style="width:480pt;"/></div>{% endif %}
  <table class="tbl">
    <tr><th class="k" style="width:15%">시각</th>{% for h in d.hours if h.hour >= 9 and h.hour < 18 %}<th>{{ h.hour }}시</th>{% endfor %}</tr>
    <tr><td class="k">내부 체감(°C)</td>{% for h in d.hours if h.hour >= 9 and h.hour < 18 %}<td style="color:{{ h.color }}; font-weight:bold;">{{ h.feels if h.feels is not none else '-' }}</td>{% endfor %}</tr>
    <tr><td class="k">외부 체감(°C)</td>{% for h in d.hours if h.hour >= 9 and h.hour < 18 %}<td style="color:#1790cd; font-weight:bold;">{{ h.out_feels if h.out_feels is not none else (h.outdoor if h.outdoor is not none else '-') }}</td>{% endfor %}</tr>
    <tr><td class="k">체감차(내-외)</td>{% for h in d.hours if h.hour >= 9 and h.hour < 18 %}<td{% if h.delta is not none and h.delta >= 5 %} style="color:#b91c1c; font-weight:bold;"{% endif %}>{{ h.delta if h.delta is not none else '-' }}</td>{% endfor %}</tr>
  </table>
  <p class="note">※ 출처: {{ '케이웨더(주)' if d.weather.provider in ('kweather', 'kma') else '참고용 추정치' }} · 외부 체감온도 = 기상청 공식 산식(측정 당시 시각 매칭, 측정기 미기록 보완값)</p>
{% elif not d.external_daily %}
  <p class="note">해당 일자의 외부 관측자료가 아직 제공되지 않아 비교 분석을 생략함.</p>
{% endif %}

<pdf:keeptogether>
<h2><span class="no">6.</span> 종합 분석</h2>
<div class="gov">{% for a in d.analysis %}<div><span class="b">□</span> {{ a }}</div>{% endfor %}</div>

<h2><span class="no">7.</span> 조치사항 및 권고 <span style="font-size:8pt; color:#64748b; font-weight:normal;">(최고 위험단계 「{{ d.peak_label }}」 기준)</span></h2>
<div class="gov2">{% for g in d.guidance %}<div><span class="b">○</span> {{ g }}</div>{% endfor %}</div>
{% else %}
<h2><span class="no">2.</span> 측정 결과</h2>
<p class="note">해당 일자에 수집된 측정 데이터가 없습니다.</p>
{% endif %}

<div class="footer">
  적용 기준: 고용노동부 「온열질환 예방가이드」(물·그늘·휴식) · 산업안전보건기준에 관한 규칙 제566조 · 폭염특보 발표 기준<br/>
  측정장비·데이터: 현장 측정값은 <b>케이웨더(주) 체감온도계 장비</b>로 측정되었으며, 외부 기상자료를 포함한 모든 데이터의 출처는 <b>케이웨더(주)</b>입니다. · 본 보고서는 케이웨더(주) 체감온도계 안전보건 대시보드에서 자동 생성되었습니다.
</div>
</pdf:keeptogether>
</body></html>
"""
)


def _html_to_pdf(html: str) -> bytes:
    buf = io.BytesIO()
    pisa.CreatePDF(src=html, dest=buf, encoding="utf-8")
    return buf.getvalue()


def daily_pdf(db: Session, tenant: Tenant, device_sn: str, on_date: date_cls, generated: str) -> bytes:
    d = _daily_detail(db, tenant, device_sn, on_date)
    report_no = f"KW-HS-{on_date.strftime('%Y%m%d')}-{str(device_sn)[-4:]}"
    chart1 = _chart_hourly_feels(d.get("series") or [], heat.thresholds()) if d.get("has_data") else None
    chart2 = _chart_compare(d.get("hours") or []) if d.get("has_data") else None
    html = _DAILY_TEMPLATE.render(
        d=d, chart=chart1, chart2=chart2, pdf_font=_PDF_FONT, generated=generated, report_no=report_no
    )
    return _html_to_pdf(html)


_PERIODIC_TEMPLATE = Template(
    """
<html><head><style>
body { font-family: "{{ pdf_font }}"; font-size: 10pt; color:#111; }
h1 { font-size: 16pt; text-align:center; border-bottom: 2px solid #0f499e; padding-bottom:6px; color:#0f172a; }
.sub { color:#475569; font-size:9pt; margin-bottom:10px; }
table { width:100%; border-collapse: collapse; margin: 8px 0; }
th, td { border:1px solid #cbd5e1; padding:4px 6px; text-align:center; }
th { background:#f1f5f9; }
.footer { color:#94a3b8; font-size:8pt; margin-top:16px; }
</style></head><body>
<h1>폭염 안전관리 기간 분석 보고서</h1>
<div class="sub">대상: {{ scope }} &nbsp;|&nbsp; 기간: {{ s.start }} ~ {{ s.end }}</div>

<table>
  <tr><th>기간 최고 체감온도</th><td>{{ s.overall_max_feels if s.overall_max_feels is not none else '-' }} °C</td>
      <th>기간 평균 체감온도</th><td>{{ s.overall_avg_feels if s.overall_avg_feels is not none else '-' }} °C</td></tr>
</table>

<h3>위험 단계 도달 일수</h3>
<table><tr>
  {% for code, lvl in levels.items() %}<th style="background:{{ lvl.color }}; color:#fff">{{ lvl.label }}</th>{% endfor %}
</tr><tr>
  {% for code in levels %}<td>{{ s.level_counts[code] }} 일</td>{% endfor %}
</tr></table>

{% if chart %}<img src="{{ chart }}" style="width:480pt;"/>{% endif %}

<h3>일자별 트렌드</h3>
<table repeat="1">
<tr><th>일자</th><th>최고 체감(°C)</th><th>평균 체감(°C)</th><th>최고온도(°C)</th><th>평균습도(%)</th><th>33°C↑(분)</th><th>단계</th></tr>
{% for row in s.daily %}
<tr><td>{{ row.date }}</td><td>{{ row.max_feels }}</td><td>{{ row.avg_feels }}</td>
<td>{{ row.max_temp }}</td><td>{{ row.avg_humidity if row.avg_humidity is not none else '-' }}</td>
<td>{{ row.minutes_over_33 }}</td><td>{{ row.peak_label }}</td></tr>
{% endfor %}
</table>
<div class="footer">자동 생성 {{ generated }}</div>
</body></html>
"""
)


def _periodic_chart(stats: dict) -> str | None:
    if not HAS_MPL or not stats["daily"]:
        return None
    days = [r["date"] for r in stats["daily"]]
    maxf = [r["max_feels"] for r in stats["daily"]]
    avgf = [r["avg_feels"] for r in stats["daily"]]
    fig, ax = plt.subplots(figsize=(9, 3.4))
    ax.plot(days, maxf, "o-", color="#dc2626", label="일 최고 체감온도")
    ax.plot(days, avgf, "o-", color="#f59e0b", label="일 평균 체감온도")
    ax.set_ylabel("체감온도 (°C)")
    ax.tick_params(axis="x", rotation=45, labelsize=7)
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.25)
    return _fig_to_data_uri(fig)


def periodic_pdf(
    db: Session, tenant: Tenant, device_sn: str | None, start: date_cls, end: date_cls, generated: str
) -> bytes:
    stats = analytics.periodic_stats(db, tenant, device_sn, start, end)
    scope = device_sn or "전체 기기"
    if device_sn:
        dev = db.get(Device, device_sn)
        if dev and dev.company_name:
            scope = f"{dev.company_name} ({device_sn})"
    html = _PERIODIC_TEMPLATE.render(
        s=stats, scope=scope, levels=heat.LEVELS, chart=_periodic_chart(stats),
        pdf_font=_PDF_FONT, generated=generated,
    )
    return _html_to_pdf(html)


# ---------------- Excel ----------------
_HEADER_FILL = PatternFill("solid", fgColor="1E293B")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

# 서버리스 응답 한도(4.5MB)·시간 제한 내에서 안전한 로우데이터 상한
EXPORT_RAW_MAX = 100_000


def export_excel(
    db: Session, tenant: Tenant, device_sn: str | None, start: datetime, end: datetime
) -> bytes:
    """Excel 내보내기 — 대용량 안전 버전.

    - 요약은 SQL 집계(전체 행을 메모리에 올리지 않음)
    - 로우데이터는 EXPORT_RAW_MAX 행으로 상한(초과 시 안내 행 추가)
    - write_only 모드로 메모리/속도 최적화
    """
    from openpyxl.cell import WriteOnlyCell
    from sqlalchemy import Date, case, cast, func

    from ..models import SensorLog

    sns = analytics._resolve_scope(db, tenant, device_sn)

    wb = Workbook(write_only=True)

    def _headers(ws, names):
        cells = []
        for n in names:
            c = WriteOnlyCell(ws, value=n)
            c.fill = _HEADER_FILL
            c.font = _HEADER_FONT
            c.alignment = Alignment(horizontal="center")
            cells.append(c)
        return cells

    # --- 시트1: 일자별 요약 (SQL 집계) ---
    ws1 = wb.create_sheet("일자별요약")
    for i, w in enumerate([12, 16, 12, 12, 12, 12, 10, 10], start=1):
        ws1.column_dimensions[chr(64 + i)].width = w
    ws1.append(_headers(ws1, ["일자", "기기SN", "최고체감(°C)", "평균체감(°C)", "최고온도(°C)", "평균습도(%)", "33°C↑(분)", "최고단계"]))

    if sns:
        date_expr = (
            func.date(SensorLog.measured_at)
            if db.bind.dialect.name == "sqlite"
            else cast(SensorLog.measured_at, Date)
        ).label("d")
        cond = [SensorLog.device_sn.in_(sns), SensorLog.measured_at >= start, SensorLog.measured_at <= end]
        q = (
            select(
                date_expr,
                SensorLog.device_sn,
                func.max(SensorLog.feels_like_temperature),
                func.avg(SensorLog.feels_like_temperature),
                func.max(SensorLog.temperature),
                func.avg(SensorLog.humidity),
                func.sum(case((SensorLog.feels_like_temperature >= settings.HEAT_CAUTION, 1), else_=0)),
            )
            .where(*cond)
            .group_by(date_expr, SensorLog.device_sn)
            .order_by(date_expr, SensorLog.device_sn)
        )
        for day, sn, mxf, avf, mxt, avh, over in db.execute(q):
            mxf = float(mxf) if mxf is not None else None
            ws1.append([
                str(day), sn,
                round(mxf, 1) if mxf is not None else None,
                round(float(avf), 1) if avf is not None else None,
                round(float(mxt), 1) if mxt is not None else None,
                round(float(avh), 1) if avh is not None else None,
                int(over or 0),
                heat.classify(mxf).label,
            ])

    # --- 시트2: 로우데이터 (상한 + 안내) ---
    ws2 = wb.create_sheet("로우데이터")
    for i, w in enumerate([22, 16, 10, 10, 12], start=1):
        ws2.column_dimensions[chr(64 + i)].width = w
    ws2.append(_headers(ws2, ["측정일시", "기기SN", "온도(°C)", "습도(%)", "체감온도(°C)"]))

    truncated = False
    if sns:
        raw_q = (
            select(
                SensorLog.measured_at, SensorLog.device_sn,
                SensorLog.temperature, SensorLog.humidity, SensorLog.feels_like_temperature,
            )
            .where(*cond)
            .order_by(SensorLog.device_sn, SensorLog.measured_at)
            .limit(EXPORT_RAW_MAX + 1)
        )
        count = 0
        for mt, sn, temp, humi, feels in db.execute(raw_q):
            count += 1
            if count > EXPORT_RAW_MAX:
                truncated = True
                break
            ws2.append([
                pd.Timestamp(mt).strftime("%Y-%m-%d %H:%M:%S"), sn,
                round(float(temp), 1) if temp is not None else None,
                int(humi) if humi is not None else None,
                round(float(feels), 1) if feels is not None else None,
            ])
    if truncated:
        note = WriteOnlyCell(ws2, value=(
            f"※ 기간 내 데이터가 {EXPORT_RAW_MAX:,}건을 초과하여 처음 {EXPORT_RAW_MAX:,}건만 수록했습니다. "
            "기간을 줄여 다시 내보내면 전체 로우데이터를 받을 수 있습니다. (일자별 요약 시트는 전체 기간 반영)"
        ))
        note.font = Font(color="DC2626", bold=True)
        ws2.append([note])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
